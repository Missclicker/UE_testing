import os
import json
import click
import subprocess
import re
import pandas as pd
import netmiko

from pathlib import Path
from datetime import datetime
from concurrent import futures
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

from config_file import *

@click.group()
def cli():
    pass


@cli.command()
@click.option('-l', '--link_check', is_flag=True, default=False, help='Gather transceiver RX info from routers')
def main(link_check):
    """check all sites availability"""
    status_st = STATUS_FILE['status']
    lbs = pd.read_excel(PATH / 'Data.xlsx', sheet_name='Loopbacks')
    vrf_ip = pd.read_excel(
        PATH / 'Data.xlsx',
        sheet_name='VRF_IPs',
        index_col=1
    ).fillna(False)

    no_cache = list()
    with futures.ThreadPoolExecutor(max_workers=20) as executor:
        [executor.submit(check_site, row, lbs, vrf_ip, no_cache) for row in status_st[2:status_st.max_row]]

    if no_cache:
        with futures.ThreadPoolExecutor(max_workers=20) as executor:
            threads = []
            print('Starting full check on sites with no cache')
            for row in no_cache:
                threads.append(executor.submit(os.system, f'site_check_mtreat.py main -f {row[0].value}'))

            for task in futures.as_completed(threads):
                _ = task.result()

        for row in no_cache:
            site_id = row[0].value
            cache_list = [
                x for x in sorted(
                    (PATH / 'json').iterdir(), key=os.path.getmtime
                ) if site_id + '_2021' in x.name
            ]
            if cache_list:
                row[4].value = cache_list[-1].stem

    name = 'site_status.xlsx'
    try:
        STATUS_FILE.save(PATH / name)
    except PermissionError:
        i = 1
        while True:
            name = f'site_status_{i}.xlsx'
            try:
                STATUS_FILE.save(PATH / name)
                break
            except PermissionError:
                print(f'File {name} is busy, increasing prefix')
                i += 1
    print(f'Saved {name}')
    if link_check:
        print('Gathering transceiver info for all avaliable sites...')
        refresh_func(debug=True)
        report_func(name)


@cli.command()
@click.option('-d', '--debug', is_flag=True, default=False, help='show ping results while refreshing')
def refresh(debug):
    return refresh_func(debug)


def refresh_func(debug=False):
    """get and save data from all routers"""
    rtrs = pd.read_excel(
        PATH / 'Data.xlsx', sheet_name='Loopbacks',
        index_col='Hostname'
    )['Loopback0'].iloc[4:]
    rtrs_data = {}
    with futures.ThreadPoolExecutor(max_workers=40) as executor:
        threads = {executor.submit(check_rtr, ip, debug): rtr for rtr, ip in rtrs.iteritems()}

        for task in futures.as_completed(threads):
            try:
                data = task.result()
            except netmiko.ssh_exception.NetmikoTimeoutException:
                print(f'Timeout on {threads[task]}')
            if data:
                rtrs_data[threads[task]] = data
            else:
                rtrs_data[threads[task]] = False
    with (PATH / f"json/{CURRENT_DATE}_fiber_link.json").open('wt') as f:
        json.dump(rtrs_data, f, indent=2)
        print(f'Saved {CURRENT_DATE}_fiber_link.json')


@cli.command()
def match_links():
    """print links for which there's no transceiver"""
    last = sorted((PATH / 'json').glob('*fiber_link.json'), key=os.path.getmtime)[-1]
    with last.open('rt') as f:
        data = json.load(f)
    links = get_links()
    for _, hosts in links[links.link_type.isin(['IS', 'DF', 'DWDM'])][
        ['Hostname A', 'Interface A', 'Hostname B', 'Interface B']
    ].iterrows():
        for name, ifs in zip(hosts[::2], hosts[1::2]):
            if 'DC' in name:
                continue
            elif data[name] and not data[name].get(ifs):
                print(f'No data for {name, ifs}')


@cli.command()
@click.argument('site_id_flag', required=False)
def report(site_id_flag):
    return report_func(site_id_flag)


def report_func(site_id_flag, name=''):
    """generate RX status for site/sites"""
    last = sorted((PATH / 'json').glob('*fiber_link.json'), key=os.path.getmtime)[-1]
    with last.open('rt') as f:
        data = json.load(f)
    links = get_links()
    # TODO filter only site in links, set filename
    if site_id_flag:
        pass
    links["ping"], links["ospf"], links["state_a"], \
        links["rx_a"], links["lthold_a"], links["hthold_a"], \
        links["state_b"], links["rx_b"], links["lthold_b"], links["hthold_b"] = \
        zip(*links.apply(lambda row: fill_link(row, data), axis=1))

    with pd.ExcelWriter(PATH / 'site_status.xlsx', engine='openpyxl', mode='a') as writer:
        wb = writer.book
        try:
            wb.remove(workBook['links'])
        except:
            pass
        links.to_excel(writer, sheet_name='links')

    status_st = STATUS_FILE['status']
    fibers = {'IS': 'G', 'DF': 'H', 'DWDM': 'I'}
    for i in range(2, status_st.max_row + 1):
        site_id = status_st[f'A{i}'].value
        for fiber in fibers.keys():
            df = links[
                ((links['Hostname A'].str.contains(f'{site_id}-RTR'))
                 | (links['Hostname B'].str.contains(f'{site_id}-RTR')))
                & (links.link_type == fiber)]
            status_st[f'{fibers[fiber]}{i}'].value, \
                status_st[f'{fibers[fiber]}{i}'].fill = fiber_summary(df)
    if not name:
        name = 'site_status.xlsx'
    STATUS_FILE.save(PATH / name)


# def df_report(site_id):
#     df = links[
#         ((links['Hostname A'].str.contains(f'{site_id}-RTR'))
#          | (links['Hostname B'].str.contains(f'{site_id}-RTR')))
#         & (links.link_type.isin(['IS', 'DF', 'DWDM']))]


def fiber_summary(df):
    if (df.state_a == df.state_b).all():
        up = df.state_a.value_counts().get('up/up', 0)
    else:
        up = '#'
    total = df.state_a.count()

    good_rx = df.apply(
        lambda row: row['lthold_a'] < row['rx_a'] < row['hthold_a'], axis=1).value_counts().get(True, 0) \
        + df.apply(lambda row: row['lthold_b'] < row['rx_b'] < row['hthold_b'], axis=1).value_counts().get(True, 0)

    ospf_full = df.ospf.value_counts().get(True, 0)

    ospf = f'{ospf_full}'
    ping = df.ping.value_counts().get('100/100', 0)
    count = f'{up}/{total}'
    rx = f'{good_rx}'
    result = f'{count} ; {rx} ; {ping} ; {ospf}'
    if up == total == ospf_full == ping:
        color = PatternFill(patternType='solid', fgColor=Color(rgb='0000FF00'))
    elif up == total:
        color = PatternFill(patternType='solid', fgColor=Color(rgb='00FFFF00'))
    else:
        color = PatternFill(patternType='solid', fgColor=Color(rgb='00FF0000'))
    return result, color


def fill_link(row, data):
    if row.link_type not in ['DF', 'DWDM', 'IS']:
        return [''] * 10
    host_a = row['Hostname A']
    int_a = row['Interface A']
    host_b = row['Hostname B']
    int_b = row['Interface B']
    if 'DC' in host_a:
        state_a = 'DC'
        rx_a = lthold_a = hthold_a = 0
        ping = data[host_b][int_b]['ping']
        ospf = data[host_b][int_b]['ospf']
        state_b = data[host_b][int_b]['state']
        rx_b = data[host_b][int_b]['lvl']['current']
        lthold_b = data[host_b][int_b]['lvl']['l_thold']
        hthold_b = data[host_b][int_b]['lvl']['h_thold']
    elif 'DC' in host_b:
        state_b = 'DC'
        rx_b = lthold_b = hthold_b = 0
        ping = data[host_a][int_a]['ping']
        ospf = data[host_a][int_a]['ospf']
        state_a = data[host_a][int_a]['state']
        rx_a = data[host_a][int_a]['lvl']['current']
        lthold_a = data[host_a][int_a]['lvl']['l_thold']
        hthold_a = data[host_a][int_a]['lvl']['h_thold']
    elif not data.get(host_a) or not data.get(host_b):
        return ['-'] * 10
    else:
        ping_b = False
        try:
            ping = data[host_a][int_a]['ping']
            ospf = data[host_a][int_a]['ospf']
            state_a = data[host_a][int_a]['state']
            rx_a = data[host_a][int_a]['lvl']['current']
            lthold_a = data[host_a][int_a]['lvl']['l_thold']
            hthold_a = data[host_a][int_a]['lvl']['h_thold']
        except:
            print(f'No data link {int_a} on {host_a}')
            state_a = 'No data'
            rx_a = lthold_a = hthold_a = 0
            ping_b = True
        try:
            if ping_b:
                ping = data[host_b][int_b]['ping']
                ospf = data[host_b][int_b]['ospf']
            state_b = data[host_b][int_b]['state']
            rx_b = data[host_b][int_b]['lvl']['current']
            lthold_b = data[host_b][int_b]['lvl']['l_thold']
            hthold_b = data[host_b][int_b]['lvl']['h_thold']
        except:
            print(f'No data link {int_b} on {host_b}')
            state_b = 'No data'
            rx_b = lthold_b = hthold_b = 0
    return ping, ospf, \
        state_a, float(rx_a), float(lthold_a), float(hthold_a), \
        state_b, float(rx_b), float(lthold_b), float(hthold_b)


def get_links():
    def link_type(row):
        if row['VLAN A']:
            return 'DG'
        elif row['Site_A'] == row['Site_B']:
            return 'IS'
        elif row['Transciever A'] in ['Integrated', '1G RJ45']:
            return 'RRL'
        elif row['Transciever A'] == '10G SR':
            return 'DWDM'
        else:
            return 'DF'

    links = pd.read_excel(
        PATH / 'Data.xlsx',
        sheet_name='Links',
        header=1,
        index_col='ID',
        keep_default_na=False
    ).iloc[18:, :17]
    site_f = lambda x: re.search('[A-Z]+-(.*)-RTR', x)[1]
    links['Site_A'] = links['Hostname A'].apply(site_f)
    links['Site_B'] = links['Hostname B'].apply(site_f)
    links['link_type'] = links.apply(lambda row: link_type(row), axis=1)
    return links


def ping_host(ip, result=True):
    reply = subprocess.run(f"ping -n 2 -w 750 {ip}", capture_output=True, text=True).stdout
    if result:
        print(reply)
    if re.search(r'Received = [1-9]', reply):
        if 'unreachable' not in reply:
            return True
    return False


def check_rtr(ip, debug):
    if ping_host(ip, False):
        if debug: print(f'Pinging {ip} OK, connecting...')
        cli_out = {}
        host = {
            'device_type': 'cisco_ios',
            'host': ip,
            'username': LOGIN,
            'password': PASSWORD,
            'secret': PASSWORD
        }
        ospf_template = PATH_TMPLT / 'cisco_ios_show_ip_ospf_neighbor.textfsm'
        int_desc_template = PATH_TMPLT / 'cisco_ios_show_interfaces_description.textfsm'
        int_trans_template = PATH_TMPLT / 'cisco_ios_show_trans.textfsm'
        with netmiko.ConnectHandler(**host, conn_timeout=10) as ssh:
            ssh.enable()
            # IFACE / RECEIVE_PWR / L_ALARM_THOLD / H_ALARM_THOLD
            cli_out['transceivers'] = ssh.send_command(
                'show int transceiver detail | b Receive',
                use_textfsm=True,
                textfsm_template=int_trans_template,
                delay_factor=2,
            )
            cli_out['descriptions'] = ssh.send_command(
                'sh int desc | include To.*RTR.*(Te|Gi).*\(.*\)',
                use_textfsm=True,
                textfsm_template=int_desc_template,
                delay_factor=2
            )
            cli_out['ospf'] = ssh.send_command(
                'show ip ospf neighbor',
                use_textfsm=True,
                textfsm_template=ospf_template
            )
            # TODO ping via links with OSPF
            ospf_ifs = {
                re.sub('[a-zA-Z]', '', i['interface']): (i['address'], i['state']) for i in cli_out['ospf']
            }
            desc_ifs = {i['port']: i for i in cli_out['descriptions']}
            result = {i['iface']: {} for i in cli_out['transceivers']}
            for i in cli_out['transceivers']:
                ifs = i['iface']

                if ifs in desc_ifs.keys():
                    result[ifs]['desc'] = desc_ifs[ifs]['descrip']
                    result[ifs]['state'] = desc_ifs[ifs]['status'] + '/' + desc_ifs[ifs]['protocol']
                else:
                    print(f'No config on {ip} {ifs}')
                    result[ifs]['desc'] = ''
                    result[ifs]['state'] = ''

                if i['iface'][2:] in ospf_ifs.keys():
                    if ospf_ifs[i['iface'][2:]][1] == 'FULL/  -':
                        result[ifs]['ospf'] = True
                    else:
                        result[ifs]['ospf'] = ospf_ifs[i['iface'][2:]][1]
                        print(result[ifs]['ospf'])
                    ping = ssh.send_command_timing(
                        f'ping {ospf_ifs[i["iface"][2:]][0]} repeat 100 size 1500',
                        delay_factor=10
                    )
                    try:
                        result[ifs]['ping'] = re.search('Success rate is [0-9]+ percent \(([0-9/]+)\)', ping)[1]
                    except:
                        print(f'no ping in {ip}')
                        result[ifs]['ping'] = ping
                else:
                    result[ifs]['ospf'] = False
                    result[ifs]['ping'] = False
                    # TODO ping from link table?

                result[ifs]['lvl'] = dict(
                    current=i['receive_pwr'],
                    h_thold=i['h_alarm_thold'],
                    l_thold=i['l_alarm_thold']
                )
        return result
    else:
        if debug: print(f'No ping from {ip}')
        return False


def ping_devs(devs):
    threads = list()
    results = list()
    with futures.ThreadPoolExecutor(max_workers=2) as executor:
        for ip in devs:
            threads.append(executor.submit(ping_host, ip, False))

        for task in futures.as_completed(threads):
            results.append(task.result())
    return results


def check_drive(site_id):
    check_string = 'S2S'
    for row in DRIVE_FILE[2:DRIVE_FILE.max_row]:
        if row[2].value == site_id and check_string not in row[14].value:
            return True


def check_site(row, lbs, vrf_ip, no_cache):
    site_id = row[0].value
    print(f'Pinging {site_id}')
    rtrs = lbs[lbs['Site ID'] == site_id]['Loopback0'].to_list()
    site_net = vrf_ip.loc[lbs[lbs['Site ID'] == site_id]['Hostname'].to_list()[0]]['Network']
    fws = [site_net[:-4] + x for x in ['12', '13']]
    if '-CP' in site_id:
        fws.pop(-1)
    rtr_ping = all(ping_devs(rtrs))
    cache_list = [
        x for x in sorted(
            (PATH / 'json').iterdir(), key=os.path.getmtime
        ) if site_id + '_2021' in x.name
    ]
    if rtr_ping:
        # os.system(f'site_check_mtreat.py main -f {site_id}')
        row[2].value = rtr_ping
        row[3].value = all(ping_devs(fws))
        if row[3].value:
            if check_drive(site_id):
                print(f'!!!!!!! FOUND ENABLED {site_id} !!!!!!!')

        if not cache_list:
            no_cache.append(row)

    if cache_list:
        row[4].value = cache_list[-1].stem

    if site_id in DRIVE_STATUS_SITES:
        row[5].value = True


if __name__ == '__main__':
    cli()
