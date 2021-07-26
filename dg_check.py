#!/usr/bin/env python
# coding: utf-8

import netmiko
from ipaddress import IPv4Address
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from ue_config import *


FAST = False


def mtu_check(ssh, vlan, ospf_d, ip):
    ping = mtu = ospf = rsvp = 'CHECK'
    if str(vlan) in ospf_d.keys() and 'FULL' in ospf_d[str(vlan)]['state']:
        # TODO rsvp check via separate TE
        rsvp = ospf = ping = True
        # rsvp = check_rsvp(ssh, ip)
    else:
        ping_cli = ssh.send_command_timing(f'ping {ip-1} count 2 timeout 1')
        if "!!" in ping_cli:
            ping = True
    if ospf != 'CHECK' and not FAST \
            or ping != 'CHECK' and ospf == 'CHECK':
        ping_cli = ssh.send_command_timing(f'ping {ip-1} count 4 timeout 1 size 1600 df')
        if "!!" in ping_cli:
            mtu = True
    elif ospf != 'CHECK':
        mtu = True
    return [ping, mtu, ospf, rsvp]


def run_check(dg, i, ssh):
    ospf_template = PATH / 'templates/cisco_ios_show_ip_ospf_neighbor.textfsm'
    ospf_cli = ssh.send_command(
        f'show ip ospf nei | i TenGigE0/0/0/24',
        use_textfsm=True,
        textfsm_template=ospf_template
    )
    ospf_d = {x['interface'].replace('TenGigE0/0/0/24.', ''): x for x in ospf_cli}
    d = (i+1) % 5
    print(f'Starting check, on DC{d} will take a long time')
    my_red = Color(rgb='00FF0000')
    my_green = Color(rgb='0000FF00')
    red_fill = PatternFill(patternType='solid', fgColor=my_red)
    green_fill = PatternFill(patternType='solid', fgColor=my_green)
    no_fill = PatternFill(fill_type=None)
    for row in dg[f'2:{dg.max_row}']:
        vlan = row[1+i].value
        ip = IPv4Address(row[2+i].value.strip())
        # print(f'Checking {vlan} on DC{d}')
        current = [x.value for x in row[3+i:7+i]]
        if current != (result := mtu_check(ssh, vlan, ospf_d, ip)):
            print(f'STATUS CHANGE ON {vlan}:\nOLD {current}\nNEW {result}')
            for cell, res in zip(row[3+i:7+i], result):
                if cell.value != res and res != 'CHECK':
                    cell.value = res
                    cell.fill = green_fill
                elif cell.value != res:
                    cell.value = res
                    cell.fill = red_fill
        else:
            # print(f'{vlan} OK')
            for cell in row[3 + i:7 + i]:
                if CURRENT_DATE not in dg.title and cell.fill == green_fill:
                    cell.fill = no_fill
    return d


def main():
    wb = load_workbook(PATH/'dg_check.xlsx', data_only=True)
    wb.copy_worksheet(wb[wb.sheetnames[-1]])
    wb[wb.sheetnames[-1]].title = f'{CURRENT_DATE}_'
    dg = wb[wb.sheetnames[-1]]
    dc1_rtr1 = {
        'device_type': 'cisco_xr',
        'host': '10.252.0.1',
        'username': LOGIN,
        'password': PASSWORD,
        'secret': PASSWORD
    }
    dc2_rtr1 = {
        'device_type': 'cisco_xr',
        'host': '10.252.0.3',
        'username': LOGIN,
        'password': PASSWORD,
        'secret': PASSWORD
    }
    print('Connecting to routers...')
    with netmiko.ConnectHandler(**dc1_rtr1) as dc1, netmiko.ConnectHandler(**dc2_rtr1) as dc2:
        print('Done!')
        with ThreadPoolExecutor(max_workers=2) as executor:
            threads = []
            for i, ssh in zip([0, 6], [dc1, dc2]):
                threads.append(executor.submit(run_check, dg, i, ssh))

            for task in as_completed(threads):
                print(f'finished on DC{task.result()}')
    wb.save(filename=PATH/'dg_check.xlsx')
    wb.close()


if __name__ == '__main__':
    main()
